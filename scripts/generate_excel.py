#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Write test cases to Excel file's "Test Case" sheet.
Creates the file if it doesn't exist, or appends to existing file.

Usage:
    python generate_excel.py <excel_file> --from-file <tsv_file>
    python generate_excel.py <excel_file> --tsv-content <base64_encoded_content>
"""

import sys
import base64
from pathlib import Path
from typing import Optional


def ensure_openpyxl():
    """Ensure openpyxl is installed"""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        return openpyxl


def write_test_cases_to_excel(
    tsv_content: str,
    excel_path: str,
    sheet_name: str = "Test Case",
    mode: str = "replace"  # "replace" or "append"
):
    """
    Write test cases to Excel file's specified sheet.
    
    Args:
        tsv_content: Tab-separated test case content
        excel_path: Path to Excel file
        sheet_name: Name of the sheet (default: "Test Case")
        mode: "replace" to clear existing content, "append" to add rows
    """
    openpyxl = ensure_openpyxl()
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')
    
    # Column widths mapping
    column_widths = {
        1: 18,   # 用例编号
        2: 15,   # 测试项
        3: 30,   # 标题
        4: 10,   # 重要级别
        5: 25,   # 预置条件
        6: 25,   # 输入
        7: 35,   # 操作步骤
        8: 30,   # 预期结果
        9: 15,   # 备注
    }
    
    excel_file = Path(excel_path)
    
    # Load or create workbook
    if excel_file.exists():
        print(f"Opening existing file: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
    else:
        print(f"Creating new file: {excel_path}")
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    
    # Get or create "Test Case" sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Found existing sheet: {sheet_name}")
        if mode == "replace":
            # Clear existing content
            ws.delete_rows(1, ws.max_row)
            print("Cleared existing content")
        else:
            # Find next available row
            start_row = ws.max_row + 1
            print(f"Appending from row {start_row}")
    else:
        ws = wb.create_sheet(title=sheet_name)
        print(f"Created new sheet: {sheet_name}")
    
    # Parse TSV content
    lines = tsv_content.strip().split('\n')
    
    # Determine start row
    if mode == "append" and ws.max_row > 0:
        # Check if header exists
        first_cell = ws.cell(row=1, column=1).value
        if first_cell:
            start_row = ws.max_row + 1
            # Skip header row when appending
            data_lines = lines[1:] if len(lines) > 1 else []
        else:
            start_row = 1
            data_lines = lines
    else:
        start_row = 1
        data_lines = lines
    
    # Write data
    for row_idx, line in enumerate(data_lines, start=start_row):
        cells = line.split('\t')
        for col_idx, cell_value in enumerate(cells, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Replace escaped newlines
            cell.value = cell_value.replace('\\n', '\n').replace('\\t', '\t')
            cell.border = border
            cell.alignment = wrap_alignment
            
            # Header row styling (row 1)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Center align for priority column (column 4)
                if col_idx == 4:
                    cell.alignment = center_alignment
    
    # Set column widths
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto filter
    if ws.max_row > 0:
        ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(excel_path)
    print(f"Successfully wrote {len(data_lines)} rows to sheet '{sheet_name}'")
    print(f"File saved: {excel_path}")
    
    return excel_path


def main():
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python generate_excel.py <excel_file> --from-file <tsv_file>")
        print("  python generate_excel.py <excel_file> --tsv-content <base64_content>")
        print("  python generate_excel.py <excel_file> --from-file <tsv_file> --sheet <sheet_name>")
        print("  python generate_excel.py <excel_file> --from-file <tsv_file> --mode append")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    sheet_name = "Test Case"
    mode = "replace"
    tsv_content = None
    
    i = 2
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == "--from-file" and i + 1 < len(sys.argv):
            tsv_file = sys.argv[i + 1]
            tsv_content = Path(tsv_file).read_text(encoding='utf-8')
            i += 2
        elif arg == "--tsv-content" and i + 1 < len(sys.argv):
            try:
                tsv_content = base64.b64decode(sys.argv[i + 1]).decode('utf-8')
            except:
                tsv_content = sys.argv[i + 1]
            i += 2
        elif arg == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]
            i += 2
        elif arg == "--mode" and i + 1 < len(sys.argv):
            mode = sys.argv[i + 1]
            i += 2
        else:
            i += 1
    
    if not tsv_content:
        print("Error: No TSV content provided")
        sys.exit(1)
    
    write_test_cases_to_excel(tsv_content, excel_file, sheet_name, mode)


if __name__ == "__main__":
    main()
